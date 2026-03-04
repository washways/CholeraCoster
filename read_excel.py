from openpyxl import load_workbook

wb = load_workbook('Unprotected MALAWI INVESTMENT TOOL 2026.xlsx')
print("Sheet names:", wb.sheetnames)
print("\n" + "="*80)

# Check each sheet for content
for sheet_name in wb.sheetnames[:5]:
    ws = wb[sheet_name]
    print(f"\nSheet: {sheet_name}")
    print(f"Dimensions: {ws.dimensions}")
    
    # Get first 30 rows with actual values
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
        # Filter to show actual content
        if any(cell is not None for cell in row[:15]):
            print(f"Row {i}: {row[:15]}")
