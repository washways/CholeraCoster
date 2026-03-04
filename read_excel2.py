from openpyxl import load_workbook

wb = load_workbook('Unprotected MALAWI INVESTMENT TOOL 2026.xlsx')

print("Sheet names:", wb.sheetnames)
print("\n" + "="*80)

# Focus on Country Model sheet
if 'Country Model' in wb.sheetnames:
    ws = wb['Country Model']
    print("\nCountry Model sheet - First 80 rows:")
    print("="*80)
    
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=80, values_only=True), 1):
        # Show rows with non-empty cells
        if any(cell is not None for cell in row[:15]):
            print(f"Row {i}: {row[:15]}")
